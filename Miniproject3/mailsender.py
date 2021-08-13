'''
아래 코드를 실행해서 NaverNewsCrawler 모듈을 임포트하세요.
# 크롤러 코드를 위한 모듈 설치
!pip install requests
!pip install beautifulsoup4
'''

from NaverNewsCrawler import NaverNewsCrawler
####사용자로 부터 기사 수집을 원하는 키워드를 input을 이용해 입력받아 ? 부분에 넣으세요
inputstr = input('검색어 : ')
if '.' in inputstr:
    inputstr = inputstr[:inputstr.index('.')]

crawler = NaverNewsCrawler(inputstr)

#### 수집한 데이터를 저장할 엑셀 파일명을 input을 이용해 입력받아 ? 부분에 넣으세요
newsfile = inputstr + '.xlsx'
crawler.get_news(newsfile)

#### 아래코드를 실행해 이메일 발송 기능에 필요한 모듈을 임포트하세요.
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
import re

#### gmail 발송 기능에 필요한 계정 정보를 아래 코드에 입력하세요.
import json 
with open ('conf.json') as f:
    config = json.load(f)

SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
SMTP_USER = config['email']
SMTP_PASSWORD = config['password']
# SMTP_USER = 'bnioinq@gmail.com'
# SMTP_PASSWORD = '1q2w3e4r'

#### 아래 코드를 실행해 메일 발송에 필요한 send_mail 함수를 만드세요.
def send_mail(name, addr, subject, contents, attachment=None):
    if not re.match('(^[a-zA-Z0-9_.-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        print('Wrong email')
        return

    msg = MIMEMultipart('alternative')
    if attachment:
        msg = MIMEMultipart('mixed')

    msg['From'] = SMTP_USER
    msg['To'] = addr
    msg['Subject'] = name + '님, ' + subject

    text = MIMEText(contents, _charset='utf-8')
    msg.attach(text)

    if attachment:
        from email.mime.base import MIMEBase
        from email import encoders

        file_data = MIMEBase('application', 'octect-stream')
        file_data.set_payload(open(attachment, 'rb').read())
        encoders.encode_base64(file_data)

        import os
        attachments = [os.path.join( os.getcwd(), 'storage', newsfile)] # 이 코드가 없으면 첨부파일이 attach(1).txt로 전송됨
        filename = os.path.basename(attachment)
        file_data.add_header('Content-Disposition', 'attachment; filename="' + filename + '"')
        msg.attach(file_data)

    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    smtp.close()

#### 프로젝트 폴더에 있는 email_list.xlsx 파일에 이메일 받을 사람들의 정보를 입력하세요.
#### 엑셀 파일의 정보를 읽어올 수 있는 모듈을 import하세요.
from openpyxl import load_workbook
wb = load_workbook('email_list.xlsx')
ws = wb.active # 엑셀워크시트 자동선택

#### email_list.xlsx 파일을 읽어와 해당 사람들에게 수집한 뉴스 정보 엑셀 파일을 send_mail 함수를 이용해 전송하세요.
i = 3
while ws.cell(i, 1).value:
    contents = inputstr +' 관련 뉴스.'
    send_mail(ws.cell(i, 2).value, ws.cell(i, 3).value, inputstr +' 관련 뉴스입니다.', contents, newsfile)
    i += 1
